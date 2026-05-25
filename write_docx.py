from docxtpl import DocxTemplate
from docxtpl import InlineImage
from docx.shared import Mm
import openpyxl

# VARIABILI GLOBALI E COSE DA IMPORTARE ============
documento_word_template = "/Users/theo/Desktop/P.IVA/Aziende/Ermes/Modelli/Modello_Relazione_RUM.docx"
OUTPUT_DOCUMENT = "/Users/theo/Desktop/Modello_Relazione_RUM_modificato.docx"

excel_data = ""

logo_formato_relazione = "" #directory logo della relazione
link_logo_relazione = "" #sito internet dell' logo della relazione
nome_azienda_generalita_relazione = "ER. Engineereing"
# ===============================================
doc = DocxTemplate(documento_word_template)

logo_azienda = "/Users/theo/Desktop/1631305251718.jpeg"
width_logo = 50

context_generalita_relazione = {
    "nome_generalita_relazione": nome_azienda_generalita_relazione,
    "logo_formato_relazione" : InlineImage(doc, logo_formato_relazione, width=Mm(20)),
    "link_logo_relazione" : InlineImage(doc, link_logo_relazione, width=Mm(5)),
    "generalita_pie_pagina" : "Via cap du cazz 89, 05479932, info@info.it"

}

context = {
    # Compila i campi vuoti
    "nome_azienda": "PARESA S.R.L.", 
    "indirizzo_azienda": "vicolo malvasia 980 (FC)",
    "note_titolo" : "relazione per la sede di mantova",
    "revisione": "rev.01",
    "data_revisione": "16/03/2026",
    "data_scadenza": "16/03/2030",
    "motivo_revisione": "Aggiornamento periodico",    #da modificare
    "img_logo_azienda" : InlineImage(doc, logo_azienda, width=Mm(width_logo)),    #da modificare
    "datore_di_lavoro" : "",    #da modificare 
    "RSPP" : "",    #da modificare
    "medico_competente": "",    #da modificare
    "RLS" : "",    #da modificare
    "giornate" : "Nelle giornate 12 e 13 settembre 2026", #modifica con le date giuste,
    #Info generali azienda
    "attivita_azienda": "",    #da modificare
    "gruppo_appartenenza" : "",    #da modificare
    "sede_legale" : "",    #da modificare
    "ubicazione_unita_operativa" : "",     #da modificare
    "date_misurazione":["05 dicembre 2025 dalle ore 08:00 alle ore 16:00",
                        "",
                        ""], #da modificare
    "strumentazione": "FUSION+DUO+FUSION+WED (01dB) / CAL 21 (01dB)",
    "condizioni_meteo": "Le condizioni meteo non hanno inficiato le misurazioni",
    "sostanze_ototossiche": "Si", #presenza di sostanze ototossiche o meno
    "misure_attuative_ototossiche": "Si faccia riferimento al documento di valutazione del rischio chimico.",
    "interazione_vib_rum": "Si", #presenza di interazione tra rumore e vibrazione
    "misure_attuative_vib_rum":"Certamente si considerato l’utilizzo di attrezzature elettriche portatili. Vi è dunque trasmissione ossea delle vibrazioni e del rumore all’orecchio medio. Si faccia riferimento alla valutazione del rischio chimico.", 
    "effetti_indesiderati": "Si",
    "misure_attuative_effetti_indesiderati": "Nelle zone/postazioni di lavoro è possibile che gli addetti possano incorrere in tali situazioni. Si consiglia pertanto di utilizzare D.P.I. con grado di protezione SNR come prescritto dalla presente relazione e l’adozione di sistemi alternativi quali segnali oto-acustici.",
    "descrizione_attivita_dettaglio":"boh", #dettaglio della mansione
}

# Lettura file Excel
wb = openpyxl.load_workbook(excel_data)

# Dizionari separati per le tabelle

# Tabella DPI: foglio "DPI art.191", righe pari (2, 4, 6, ...) fino a riga vuota
dpi_sheet = wb["DPI art.191"]
_dpi_rows = []
_row = 2
while True:
    if dpi_sheet.cell(row=_row, column=1).value is None:
        break
    _dpi_rows.append({
        "codice_DPI": str(dpi_sheet.cell(row=_row, column=1).value or ""),
        "descrizione": str(dpi_sheet.cell(row=_row, column=2).value or ""),
        "marca":       str(dpi_sheet.cell(row=_row, column=3).value or ""),
        "modello":     str(dpi_sheet.cell(row=_row, column=4).value or ""),
        "snr":         str(dpi_sheet.cell(row=_row, column=5).value or ""),
        "H":           str(dpi_sheet.cell(row=_row, column=6).value or ""),
        "L":           str(dpi_sheet.cell(row=_row, column=7).value or ""),
        "M":           str(dpi_sheet.cell(row=_row, column=8).value or ""),
        "note":        "",
    })
    _row += 2

context_tabella_dpi = {"tabella_dpi": _dpi_rows}

context_tabella_orari = {
    "tabella_orario_lavoro_mansione":[
        {"mansione": "Nome mansione", #da modificare
        "orario_lavoro" : "Lunedì – Venerdì \n 8:00÷12:00 13:00÷17:00" #da modificare
        }
    ],
}

# Tabella mansioni: fogli "Scheda 1", "Scheda 2", ... — ID=nome foglio, Mansione=F8, N_lavoratori=1
_mansioni_rows = []
_n = 1
while True:
    _sheet_name = f"Scheda {_n}"
    if _sheet_name not in wb.sheetnames:
        break
    _s = wb[_sheet_name]
    _mansioni_rows.append({
        "ID":           _sheet_name,
        "Mansione":     str(_s["F8"].value or ""),
        "N_lavoratori": 1,
    })
    _n += 1

context_tabella_mansioni = {"tabella_mansioni_numero_lavoratori": _mansioni_rows}

# Tabella HEG: foglio "Riepilogo", da riga 2 fino a riga vuota — codice_HEG generato come M01, M02, ...
_heg_sheet = wb["Riepilogo"]
_heg_rows = []
_row = 2
_idx = 1
while True:
    if _heg_sheet.cell(row=_row, column=1).value is None:
        break
    _par = _heg_sheet.cell(row=_row, column=3).value
    _heg_rows.append({
        "gruppo_HEG":           str(_heg_sheet.cell(row=_row, column=1).value or ""),
        "numero_scheda":        str(_heg_sheet.cell(row=_row, column=2).value or ""),
        "codice_HEG":           f"M{_idx:02d}",
        "parametro_riferimento": str(_par) if _par is not None else "Lex,8h",
        "lex8h":                str(_heg_sheet.cell(row=_row, column=4).value or ""),
        "incertezza":           str(_heg_sheet.cell(row=_row, column=5).value or ""),
        "lexmax":               str(_heg_sheet.cell(row=_row, column=6).value or ""),
        "peakmax":              str(_heg_sheet.cell(row=_row, column=7).value or ""),
        "classe_rischio":       str(_heg_sheet.cell(row=_row, column=10).value or ""),
        "esposizione_vibrazioni": str(_heg_sheet.cell(row=_row, column=11).value or ""),
        "esposizione_ototossici": str(_heg_sheet.cell(row=_row, column=12).value or ""),
        "rumori_impulsivi":     str(_heg_sheet.cell(row=_row, column=13).value or ""),
    })
    _row += 1
    _idx += 1

context_tabella_heg = {"tabella_HEG": _heg_rows}

# Unione di tutti i dizionari
context_completo = context_generalita_relazione | context | context_tabella_dpi | context_tabella_orari | context_tabella_mansioni | context_tabella_heg

doc.render(context_completo)
doc.save(OUTPUT_DOCUMENT)
print('Docx scritto')