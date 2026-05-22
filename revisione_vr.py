"""
revisione_vr.py
Automazione del flusso di lavoro per la revisione della valutazione del rischio rumore.

Flusso principale (orchestratore):
  1. Copia i dati dai file data-excel nel vr-excel
  2. Salva il vr-excel nella directory di output
  3. Pausa – l'utente modifica manualmente il file (Schede HEG, DPI)
  4. Ricarica, nasconde righe, imposta aree di stampa, salva
  5. Esporta i PDF di Tab-Mis e di tutte le Schede
"""

import os
import re
import shutil
import logging
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

import config

# =============================================================================
# Logging
# =============================================================================

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(
            os.path.join(os.path.dirname(__file__), "revisione_vr.log"),
            mode="w",
            encoding="utf-8",
        ),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# Utility helpers
# =============================================================================

def col(letter: str) -> int:
    """Converte una lettera di colonna Excel (es. 'AA') in indice 1-based."""
    return column_index_from_string(letter)


def find_last_row(ws, col_letter: str) -> int:
    """
    Ritorna l'ultimo numero di riga non vuota nella colonna specificata.
    Usa ws._cells (stato in-memory) invece di ws.max_row per evitare
    problemi con l'attributo <dimension> stale nei file xlsx caricati.
    """
    col_idx = col(col_letter)
    rows_in_col = [r for (r, c) in ws._cells if c == col_idx]
    return max(rows_in_col) if rows_in_col else 0


def adjust_formula_row(formula: str, row_offset: int) -> str:
    """
    Aggiusta i riferimenti di riga relativi in una formula Excel.
    I riferimenti assoluti (es. $A$1) non vengono modificati.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _replace(match):
        dollar_col = match.group(1)
        col_letters = match.group(2)
        dollar_row = match.group(3)
        row_str = match.group(4)
        if dollar_row:          # riferimento assoluto di riga → invariato
            return match.group(0)
        new_row = int(row_str) + row_offset
        return f"{dollar_col}{col_letters}{new_row}"

    return re.sub(r"(\$?)([A-Za-z]+)(\$?)(\d+)", _replace, formula)


def get_schede_sheets(wb) -> list:
    """Ritorna la lista ordinata dei nomi dei fogli 'Scheda N' del workbook."""
    pattern = re.compile(r"^Scheda\s+\d+$", re.IGNORECASE)
    sheets = [ws.title for ws in wb.worksheets if pattern.match(ws.title)]
    # Ordina per numero (es. Scheda 1, Scheda 2, ...)
    sheets.sort(key=lambda s: int(re.search(r"\d+", s).group()))
    return sheets


def export_sheet_to_pdf(wb_path: str, sheet_name: str, output_pdf_path: str):
    """
    Esporta un singolo foglio in PDF tramite LibreOffice CLI.
    Crea una copia temporanea del workbook con gli altri fogli nascosti,
    la converte con LibreOffice e sposta il PDF nella posizione finale.

    Prerequisito: LibreOffice deve essere installato e nel PATH
    (su macOS tipicamente in /Applications/LibreOffice.app/Contents/MacOS/soffice).
    """
    # Cerca il binario libreoffice / soffice
    lo_bin = _find_libreoffice()
    if lo_bin is None:
        raise EnvironmentError(
            "LibreOffice non trovato. Installalo o aggiungi il percorso a PATH."
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_xlsx = os.path.join(tmpdir, "temp_export.xlsx")

        wb_temp = openpyxl.load_workbook(wb_path)
        for ws in wb_temp.worksheets:
            ws.sheet_state = "visible" if ws.title == sheet_name else "hidden"

        wb_temp.save(tmp_xlsx)

        cmd = [lo_bin, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, tmp_xlsx]
        log.debug("LibreOffice cmd: %s", " ".join(cmd))
        result = subprocess.run(cmd, capture_output=True, text=True)

        if result.returncode != 0:
            log.error("LibreOffice stderr: %s", result.stderr)
            raise RuntimeError(
                f"Errore esportazione PDF foglio '{sheet_name}': {result.stderr}"
            )

        tmp_pdf = os.path.join(tmpdir, "temp_export.pdf")
        if not os.path.exists(tmp_pdf):
            raise FileNotFoundError(f"PDF atteso non trovato: {tmp_pdf}")

        os.makedirs(os.path.dirname(output_pdf_path), exist_ok=True)
        shutil.move(tmp_pdf, output_pdf_path)
        log.info("PDF esportato: %s", output_pdf_path)


def _find_libreoffice() -> str | None:
    """Cerca il binario di LibreOffice nel PATH e nei percorsi comuni macOS."""
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    # macOS default install
    macos_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(macos_path):
        return macos_path
    return None


# =============================================================================
# Funzione 1 – data_excel2vr_excel
# =============================================================================

def data_excel2vr_excel(
    path_data_excel: str,
    path_vr_excel: str,
    revisione_numero: int,
    data_misure: str,
    nome_foglio: str = "Tab-Mis",
    wb_vr=None,
) -> openpyxl.Workbook:
    """
    Legge un file data-excel (mis*.xlsx) e copia i dati nel foglio nome_foglio
    del file vr-excel secondo le specifiche del documento funzionale.

    Parametri colonne nel data-excel (1-indexed):
      B=2  letter_ID,  C=3  nTrack,  F=6  LeqA_eq,  I=9  LeqC_eq,  J=10 PeakC_max

    Parametri colonne target in vr-excel (Tab-Mis):
      ID:     colonna B
      300:    colonne Q, W, AA, AE, AI, AM
      Track1: R, S, T
      Track2: X, Y, Z
      Track3: AB, AC, AD
      Track4: AF, AG, AH
      Track5: AJ, AK, AL
      Track6: AN, AO, AP
      Formule: M, N, O, P  (copiate dall'ultima riga template)

    Se wb_vr è None il workbook viene caricato da disco e salvato alla fine.
    Se wb_vr è già caricato (passato da orchestrazione_copia_incolla) viene
    usato direttamente e il salvataggio è delegato al chiamante.
    Ritorna sempre il workbook modificato.
    """
    # --- Configurazione colonne data-excel ---
    DE_COL_LETTER_ID = "B"
    DE_COL_NTRACK    = "C"
    DE_COL_F         = "F"   # LeqA_eq
    DE_COL_I         = "I"   # LeqC_eq
    DE_COL_J         = "J"   # PeakC_max
    DE_HEADER_ROW    = 1

    # --- Configurazione colonne vr-excel ---
    VR_COL_ID         = "B" # prima era B
    VR_COLS_300       = ["Q", "W", "AA", "AE", "AI", "AM"]
    VR_FORMULA_COLS   = ["M", "N", "O", "P"]
    VR_NTRACK_MAP     = {
        1: ("R",  "S",  "T"),
        2: ("X",  "Y",  "Z"),
        3: ("AB", "AC", "AD"),
        4: ("AF", "AG", "AH"),
        5: ("AJ", "AK", "AL"),
        6: ("AN", "AO", "AP"),
    }

    log.info("=== data_excel2vr_excel ===")
    log.info("data-excel: %s", path_data_excel)
    log.info("vr-excel:   %s", path_vr_excel)
    log.info("revisione:  %d  |  data: %s  |  foglio: %s", revisione_numero, data_misure, nome_foglio)

    # --- Carica data-excel ---
    log.debug("Caricamento data-excel...")
    wb_data = openpyxl.load_workbook(path_data_excel, data_only=True)
    ws_data = wb_data.active

    # Raccogli: {letter_id: {ntrack: (leqa_eq, leqc_eq, peakc_max)}}
    data_per_id: dict[str, dict[int, tuple]] = {}
    for row in ws_data.iter_rows(min_row=DE_HEADER_ROW + 1, values_only=True):
        letter_id = row[col(DE_COL_LETTER_ID) - 1]
        ntrack    = row[col(DE_COL_NTRACK) - 1]
        f_val     = row[col(DE_COL_F) - 1]
        i_val     = row[col(DE_COL_I) - 1]
        j_val     = row[col(DE_COL_J) - 1]

        if letter_id is None:
            continue

        letter_id = str(letter_id).strip()
        if letter_id not in data_per_id:
            data_per_id[letter_id] = {}

        if ntrack is not None:
            # Arrotonda a 1 decimale
            data_per_id[letter_id][int(ntrack)] = (
                round(float(f_val), 1) if f_val is not None else None,
                round(float(i_val), 1) if i_val is not None else None,
                round(float(j_val), 1) if j_val is not None else None,
            )

    unique_ids = list(data_per_id.keys())
    log.info("ID univoci trovati nel data-excel: %s", unique_ids)

    # --- Carica vr-excel (se non già passato dal chiamante) ---
    _owner = wb_vr is None   # True = questa funzione ha aperto il wb e deve salvarlo
    if _owner:
        log.debug("Caricamento vr-excel da disco...")
        wb_vr = openpyxl.load_workbook(path_vr_excel)
    else:
        log.debug("Uso workbook vr-excel già caricato in memoria.")

    if nome_foglio not in wb_vr.sheetnames:
        raise ValueError(f"Foglio '{nome_foglio}' non trovato nel vr-excel.")

    ws_vr = wb_vr[nome_foglio]

    # Trova l'ultima riga occupata in colonna B (per sapere dove inserire)
    last_row_b = find_last_row(ws_vr, "F")
    log.info("Ultima riga occupata in colonna B del foglio '%s': %d", nome_foglio, last_row_b)

    # Individua riga template per le formule (ultima riga con dati in col B)
    template_row = last_row_b
    if template_row < 2:
        log.warning(
            "Nessuna riga template trovata per le formule (col B vuota o solo header). "
            "Le formule M/N/O/P non saranno copiate."
        )
        template_row = None

    # --- Inserisci i dati riga per riga ---
    for letter_id in unique_ids:
        new_row = last_row_b + 1
        last_row_b = new_row

        id_with_rev = f"{letter_id}_{revisione_numero}"
        log.debug("Inserimento riga %d: ID=%s", new_row, id_with_rev)

        # Colonna B: ID con pedice revisione
        ws_vr.cell(row=new_row, column=col(VR_COL_ID)).value = id_with_rev

        # Colonne con valore fisso 300
        for c_letter in VR_COLS_300:
            ws_vr.cell(row=new_row, column=col(c_letter)).value = 300

        # Copia valori per ogni nTrack
        ntrack_data = data_per_id[letter_id]
        for ntrack, (f_val, i_val, j_val) in ntrack_data.items():
            if ntrack not in VR_NTRACK_MAP:
                log.warning("nTrack=%d non mappato per ID=%s, ignorato.", ntrack, letter_id)
                continue
            col_f_target, col_i_target, col_j_target = VR_NTRACK_MAP[ntrack]
            ws_vr.cell(row=new_row, column=col(col_f_target)).value = f_val
            ws_vr.cell(row=new_row, column=col(col_i_target)).value = i_val
            ws_vr.cell(row=new_row, column=col(col_j_target)).value = j_val
            log.debug(
                "  nTrack=%d → %s=%s, %s=%s, %s=%s",
                ntrack, col_f_target, f_val, col_i_target, i_val, col_j_target, j_val,
            )

        # Copia formule da riga template (M, N, O, P)
        if template_row is not None:
            row_offset = new_row - template_row
            for formula_col in VR_FORMULA_COLS:
                template_cell = ws_vr.cell(row=template_row, column=col(formula_col))
                template_val  = template_cell.value
                adjusted = adjust_formula_row(str(template_val), row_offset) if template_val else None
                ws_vr.cell(row=new_row, column=col(formula_col)).value = adjusted
                log.debug("  Formula col %s: '%s' → '%s'", formula_col, template_val, adjusted)

    if _owner:
        log.info("Dati inseriti. Salvataggio vr-excel in: %s", path_vr_excel)
        wb_vr.save(path_vr_excel)
    else:
        log.debug("Dati inseriti in memoria (salvataggio delegato al chiamante).")

    log.info("=== data_excel2vr_excel completata ===")
    return wb_vr


# =============================================================================
# Funzione 2 – orchestrazione_copia_incolla
# =============================================================================

def orchestrazione_copia_incolla(path_data_excel: str, path_vr_excel: str) -> None:
    """
    Itera su tutti i file mis*.xlsx nella directory path_data_excel e
    chiama data_excel2vr_excel per ciascuno.
    """
    log.info("=== orchestrazione_copia_incolla ===")
    log.info("Directory data-excel: %s", path_data_excel)

    data_files = sorted(
        f for f in os.listdir(path_data_excel)
        if f.lower().startswith("mis") and f.lower().endswith(".xlsx")
    )

    if not data_files:
        log.warning("Nessun file mis*.xlsx trovato in: %s", path_data_excel)
        return

    log.info("File trovati: %s", data_files)

    # Carica il vr-excel UNA SOLA VOLTA: evita problemi con l'attributo
    # <dimension> stale che farebbe restituire sempre lo stesso max_row
    # a ogni ricarica dal disco dopo un salvataggio parziale.
    log.debug("Caricamento vr-excel una sola volta: %s", path_vr_excel)
    wb_vr = openpyxl.load_workbook(path_vr_excel)

    for filename in data_files:
        full_path = os.path.join(path_data_excel, filename)
        log.info("--- Elaborazione: %s ---", filename)
        wb_vr = data_excel2vr_excel(
            path_data_excel=full_path,
            path_vr_excel=path_vr_excel,
            revisione_numero=config.REVISIONE_NUMERO,
            data_misure=config.DATA_MISURE,
            wb_vr=wb_vr,
        )

    # Salva il workbook una sola volta alla fine
    log.info("Salvataggio vr-excel: %s", path_vr_excel)
    wb_vr.save(path_vr_excel)
    log.info("=== orchestrazione_copia_incolla completata ===")


# =============================================================================
# Funzione 3 – nascondi_righe_tab_mis
# =============================================================================

def nascondi_righe_tab_mis(
    wb,
    nome_foglio: str = "Tab-Mis",
) -> None:
    """
    Raccoglie i valori della colonna E (da riga 11 in poi, fino alla prima
    riga nascosta) da tutti i fogli 'Scheda N', poi nasconde le righe del
    foglio nome_foglio dove la colonna B è presente in quella lista.
    """
    # --- Configurazione ---
    SCHEDA_COL_E          = "E"
    SCHEDA_START_ROW      = 11
    TAB_MIS_COL_B         = "B"
    TAB_MIS_DATA_START    = 2   # riga da cui inizia i dati (salta header)

    log.info("=== nascondi_righe_tab_mis ===")

    schede = get_schede_sheets(wb)
    if not schede:
        log.warning("Nessun foglio 'Scheda N' trovato nel workbook.")
        return

    # Raccoglie valori da colonna E di tutte le schede
    id_list: set = set()
    for sheet_name in schede:
        ws = wb[sheet_name]
        log.debug("Lettura colonna E dal foglio '%s' (da riga %d)", sheet_name, SCHEDA_START_ROW)
        for row_num in range(SCHEDA_START_ROW, ws.max_row + 1):
            # Fermati alla prima riga nascosta
            if ws.row_dimensions[row_num].hidden:
                log.debug("  Riga %d nascosta, stop.", row_num)
                break
            cell_val = ws.cell(row=row_num, column=col(SCHEDA_COL_E)).value
            if cell_val is not None and str(cell_val).strip():
                id_list.add(str(cell_val).strip())

    log.info("ID raccolti dalle Schede: %s", sorted(id_list))

    # Nasconde le righe di Tab-Mis i cui ID NON sono presenti nelle Schede
    if nome_foglio not in wb.sheetnames:
        raise ValueError(f"Foglio '{nome_foglio}' non trovato nel workbook.")

    ws_tab = wb[nome_foglio]
    hidden_count = 0
    for row_num in range(TAB_MIS_DATA_START, ws_tab.max_row + 1):
        cell_b = ws_tab.cell(row=row_num, column=col(TAB_MIS_COL_B)).value
        if cell_b is None or str(cell_b).strip() not in id_list:
            ws_tab.row_dimensions[row_num].hidden = True
            hidden_count += 1
            log.debug("  Riga %d nascosta (ID=%s non in lista)", row_num, cell_b)

    log.info("Righe nascoste in '%s': %d", nome_foglio, hidden_count)
    log.info("=== nascondi_righe_tab_mis completata ===")


# =============================================================================
# Funzione 4 – select_area_stampa
# =============================================================================

def select_area_stampa(wb) -> None:
    """
    Imposta le aree di stampa:
      - Foglio Tab-Mis:    A1:BC<ultima_riga_popolata>
      - Fogli Scheda N:    A1:Q66
    """
    # --- Configurazione ---
    TAB_MIS_PRINT_START  = "A1"
    TAB_MIS_PRINT_END_COL = "BC"
    TAB_MIS_SEARCH_COL   = "B"   # usata per trovare l'ultima riga
    SCHEDA_PRINT_AREA    = "A1:Q66"

    log.info("=== select_area_stampa ===")

    # Tab-Mis
    if "Tab-Mis" in wb.sheetnames:
        ws_tab = wb["Tab-Mis"]
        last_row = find_last_row(ws_tab, TAB_MIS_SEARCH_COL)
        area = f"{TAB_MIS_PRINT_START}:{TAB_MIS_PRINT_END_COL}{last_row}"
        ws_tab.print_area = area
        log.info("Area stampa Tab-Mis: %s", area)
    else:
        log.warning("Foglio 'Tab-Mis' non trovato, skip area stampa.")

    # Fogli Scheda N
    schede = get_schede_sheets(wb)
    for sheet_name in schede:
        wb[sheet_name].print_area = SCHEDA_PRINT_AREA
        log.info("Area stampa '%s': %s", sheet_name, SCHEDA_PRINT_AREA)

    log.info("=== select_area_stampa completata ===")


# =============================================================================
# Funzione 5 – esporta_pdf_misure
# =============================================================================

def esporta_pdf_misure(
    wb_path: str,
    output_dir: str,
    output_name: str = "Tabella_Misure.pdf",
) -> None:
    """Esporta in PDF il foglio Tab-Mis del file wb_path."""
    log.info("=== esporta_pdf_misure ===")
    select_area_stampa(openpyxl.load_workbook(wb_path))  # riapplica per sicurezza

    # Riapplica + salva prima di esportare
    wb = openpyxl.load_workbook(wb_path)
    select_area_stampa(wb)
    wb.save(wb_path)

    output_pdf = os.path.join(output_dir, output_name)
    export_sheet_to_pdf(wb_path, "Tab-Mis", output_pdf)
    log.info("=== esporta_pdf_misure completata: %s ===", output_pdf)


# =============================================================================
# Funzione 6 – esporta_pdf_schede
# =============================================================================

def esporta_pdf_schede(wb_path: str, output_dir: str) -> None:
    """
    Esporta in PDF tutti i fogli 'Scheda N' per ogni valore DPI della tabella
    'DPI art.191' (colonna A, righe 2, 4, 6, ...).

    Nome output: <nome_scheda>_DPI_<id_DPI>.pdf
    dove id_DPI è letto dalla cella E62 di ogni foglio Scheda.

    Configurazione colonne DPI art.191:
      - Colonna A: ID DPI (righe pari: 2, 4, 6, ...)
    """
    # --- Configurazione ---
    FOGLIO_DPI         = "DPI art.191"
    DPI_COL_ID         = "A"
    DPI_START_ROW      = 2
    DPI_STEP           = 2    # righe pari
    SCHEDA_DPI_CELL    = "E62"

    log.info("=== esporta_pdf_schede ===")

    wb = openpyxl.load_workbook(wb_path)

    if FOGLIO_DPI not in wb.sheetnames:
        raise ValueError(f"Foglio '{FOGLIO_DPI}' non trovato nel workbook.")

    ws_dpi = wb[FOGLIO_DPI]
    schede = get_schede_sheets(wb)

    if not schede:
        log.warning("Nessun foglio 'Scheda N' trovato. Esportazione interrotta.")
        return

    # Raccoglie gli ID DPI (righe 2, 4, 6, ...)
    dpi_ids = []
    row_num = DPI_START_ROW
    while True:
        val = ws_dpi.cell(row=row_num, column=col(DPI_COL_ID)).value
        if val is None:
            break
        dpi_ids.append(str(val).strip())
        row_num += DPI_STEP

    log.info("ID DPI trovati: %s", dpi_ids)

    os.makedirs(output_dir, exist_ok=True)

    for dpi_id in dpi_ids:
        log.info("--- DPI: %s ---", dpi_id)

        # Imposta E62 in tutte le schede e salva
        for sheet_name in schede:
            ws = wb[sheet_name]
            ws[SCHEDA_DPI_CELL] = dpi_id

        # Riapplica aree di stampa e salva in un file temporaneo
        select_area_stampa(wb)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_f:
            tmp_path = tmp_f.name
        try:
            wb.save(tmp_path)

            # Esporta ogni scheda
            for sheet_name in schede:
                # id_DPI letto dalla cella E62 della scheda stessa
                actual_dpi = wb[sheet_name][SCHEDA_DPI_CELL].value or dpi_id
                safe_sheet = re.sub(r"[^\w\s-]", "", sheet_name).strip().replace(" ", "_")
                safe_dpi   = re.sub(r"[^\w-]", "", str(actual_dpi))
                pdf_name   = f"{safe_sheet}_DPI_{safe_dpi}.pdf"
                output_pdf = os.path.join(output_dir, pdf_name)

                log.info("Esportazione: %s", pdf_name)
                export_sheet_to_pdf(tmp_path, sheet_name, output_pdf)

        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    log.info("=== esporta_pdf_schede completata ===")


# =============================================================================
# Funzione 7 – orchestratore
# =============================================================================

def orchestratore(
    path_output: str = config.PATH_OUTPUT,
    nome_vr_excel_out: str = config.NOME_VR_EXCEL_OUT,
    skip_to_post_pausa: bool = config.SKIP_TO_POST_PAUSA,
) -> None:
    """
    Flusso di lavoro principale.

    Se skip_to_post_pausa=False (default):
      1. Copia dati dai data-excel nel vr-excel
      2. Salva il vr-excel in PATH_OUTPUT
      3. Pausa: attende che l'utente modifichi il file

    Poi (sempre):
      4. Ricarica il file
      5. Nasconde righe in Tab-Mis
      6. Imposta aree di stampa
      7. Salva
      8. Esporta PDF Tab-Mis
      9. Esporta PDF Schede
    """
    log.info("========== ORCHESTRATORE ==========")
    log.info("OUTPUT DIR:  %s", path_output)
    log.info("OUTPUT FILE: %s", nome_vr_excel_out)
    log.info("SKIP FASE 1: %s", skip_to_post_pausa)

    os.makedirs(path_output, exist_ok=True)
    pdf_dir = os.path.join(path_output, "PDFs")
    os.makedirs(pdf_dir, exist_ok=True)

    output_vr_path = os.path.join(path_output, nome_vr_excel_out)

    # ---------------------------------------------------------------
    # FASE 1: Copia dati (opzionale)
    # ---------------------------------------------------------------
    if not skip_to_post_pausa:
        log.info("--- FASE 1: copia dati ---")
        # Copia il template nella directory di output prima di modificarlo
        shutil.copy2(config.PATH_VR_EXCEL, output_vr_path)
        log.info("Template copiato in: %s", output_vr_path)

        orchestrazione_copia_incolla(
            path_data_excel=config.PATH_DATA_EXCEL,
            path_vr_excel=output_vr_path,
        )

        log.info("--- FASE 1 completata, vr-excel salvato in: %s ---", output_vr_path)

        # Pausa
        print("\n" + "=" * 60)
        print(f"In attesa di modifica del documento '{nome_vr_excel_out}'")
        print("Modifica le Schede HEG e i DPI del documento.")
        print(f"File: {output_vr_path}")
        print("=" * 60)
        input("Premi INVIO per continuare...")
    else:
        log.info("--- SKIP FASE 1: si parte dal file esistente: %s ---", output_vr_path)
        if not os.path.exists(output_vr_path):
            raise FileNotFoundError(
                f"SKIP_TO_POST_PAUSA=True ma il file non esiste: {output_vr_path}"
            )

    # ---------------------------------------------------------------
    # FASE 2: Post-pausa
    # ---------------------------------------------------------------
    log.info("--- FASE 2: post-pausa ---")

    # 4. Ricarica
    log.info("Ricaricamento workbook: %s", output_vr_path)
    wb = openpyxl.load_workbook(output_vr_path)

    # 5. Nascondi righe
    nascondi_righe_tab_mis(wb)

    # 6. Aree di stampa
    select_area_stampa(wb)

    # 7. Salva
    wb.save(output_vr_path)
    log.info("Workbook salvato: %s", output_vr_path)

    # 8. Esporta PDF Tab-Mis
    esporta_pdf_misure(
        wb_path=output_vr_path,
        output_dir=pdf_dir,
        output_name=config.NOME_PDF_TABELLA_MISURE,
    )

    # 9. Esporta PDF Schede
    esporta_pdf_schede(
        wb_path=output_vr_path,
        output_dir=pdf_dir,
    )

    log.info("========== ORCHESTRATORE COMPLETATO ==========")
    print(f"\nCompletato. Output in: {path_output}")


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    orchestratore()
